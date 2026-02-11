"""
报告导出服务
支持导出PDF和Excel格式的核对报告
"""

import json
import os
import platform
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, Image, KeepTogether
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT


def find_and_register_font():
    """查找并注册系统中可用的中文字体"""
    # 首先尝试使用项目自带的字体
    base_dir = Path(__file__).parent.parent
    local_font = base_dir / 'fonts' / 'NotoSansCJKsc-Regular.otf'

    if local_font.exists():
        try:
            font_name = 'NotoSansCJKsc'
            pdfmetrics.registerFont(TTFont(font_name, str(local_font)))
            print(f"[INFO] PDF导出服务使用本地字体: {font_name}")
            return font_name
        except Exception as e:
            print(f"[WARN] 本地字体加载失败: {e}")

    # 如果没有本地字体，尝试系统字体
    system = platform.system()
    font_candidates = []

    if system == 'Darwin':  # macOS
        # 优先使用支持中文的 TrueType 字体
        font_candidates = [
            ('/Library/Fonts/Arial Unicode.ttf', 'ArialUnicode'),
            ('/System/Library/Fonts/Supplemental/Arial Unicode.ttf', 'ArialUnicode'),
        ]
    elif system == 'Windows':
        font_candidates = [
            ('C:/Windows/Fonts/simhei.ttf', 'SimHei'),
            ('C:/Windows/Fonts/simsun.ttc', 'SimSun'),
            ('C:/Windows/Fonts/msyh.ttc', 'MSYaHei'),
        ]
    else:  # Linux
        font_candidates = [
            ('/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc', 'WenQuanYi'),
        ]

    for font_path, font_name in font_candidates:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                print(f"[INFO] PDF导出服务使用字体: {font_name} ({font_path})")
                return font_name
            except Exception as e:
                print(f"[WARN] 字体加载失败: {font_path} - {e}")
                continue

    print("[WARN] 未找到中文字体，PDF中文可能显示为方块")
    return 'Helvetica'


# 全局字体名称
FONT_NAME = find_and_register_font()


def to_para(text, style):
    """将文本转换为Paragraph，确保字体正确"""
    if text is None:
        text = ''
    else:
        text = str(text)
    return Paragraph(text, style)


class ReportExportService:
    """报告导出服务"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_styles()

    def _setup_styles(self):
        """设置PDF样式"""
        # 标题样式
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontName=FONT_NAME,
            fontSize=20,
            alignment=TA_CENTER,
            spaceAfter=20,
        )

        # 章节标题
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontName=FONT_NAME,
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
        )

        # 小标题
        self.subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=self.styles['Heading3'],
            fontName=FONT_NAME,
            fontSize=12,
            spaceBefore=10,
            spaceAfter=5,
        )

        # 正文
        self.body_style = ParagraphStyle(
            'CustomBody',
            parent=self.styles['BodyText'],
            fontName=FONT_NAME,
            fontSize=10,
            spaceBefore=3,
            spaceAfter=3,
        )

        # 说明文字
        self.note_style = ParagraphStyle(
            'NoteStyle',
            parent=self.styles['BodyText'],
            fontName=FONT_NAME,
            fontSize=9,
            textColor=colors.grey,
            spaceBefore=2,
            spaceAfter=2,
        )

        # 表格单元格样式
        self.cell_style = ParagraphStyle(
            'CellStyle',
            parent=self.styles['BodyText'],
            fontName=FONT_NAME,
            fontSize=9,
            leading=12,
        )

    def export_pdf(self, result: Dict[str, Any], output_path: str) -> str:
        """导出PDF报告"""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
        )

        elements = []

        # 1. 封面/标题
        elements.extend(self._create_header(result))

        # 2. 统计概览
        elements.extend(self._create_statistics(result))

        # 3. 首页与第三页比对
        elements.extend(self._create_home_third_comparison(result))

        # 4. 部件核对详情
        elements.extend(self._create_component_details(result))

        # 5. 问题汇总
        elements.extend(self._create_issues_summary(result))

        # 6. 页脚信息
        elements.extend(self._create_footer(result))

        # 生成PDF
        doc.build(elements)

        return output_path

    def export_excel(self, result: Dict[str, Any], output_path: str) -> str:
        """导出Excel报告"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()

        # 1. 概览sheet
        ws_overview = wb.active
        ws_overview.title = "核对概览"
        self._fill_excel_overview(ws_overview, result)

        # 2. 部件核对sheet
        ws_components = wb.create_sheet("部件核对")
        self._fill_excel_components(ws_components, result)

        # 3. 问题汇总sheet
        ws_issues = wb.create_sheet("问题汇总")
        self._fill_excel_issues(ws_issues, result)

        wb.save(output_path)
        return output_path

    def _create_header(self, result: Dict[str, Any]) -> List:
        """创建报告标题"""
        elements = []

        elements.append(Paragraph("PDF报告核对结果", self.title_style))
        elements.append(Spacer(1, 10))

        filename = result.get('filename', '未知文件')
        check_time = result.get('check_time', '')
        file_id = result.get('file_id', '')

        info_data = [
            [to_para('文件名', self.cell_style), to_para(filename, self.cell_style)],
            [to_para('核对时间', self.cell_style), to_para(check_time, self.cell_style)],
            [to_para('报告ID', self.cell_style), to_para(file_id[:8] + '...', self.cell_style)],
        ]

        info_table = Table(info_data, colWidths=[4*cm, 12*cm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ]))

        elements.append(info_table)
        elements.append(Spacer(1, 20))

        return elements

    def _create_statistics(self, result: Dict[str, Any]) -> List:
        """创建统计概览"""
        elements = []

        elements.append(Paragraph("一、核对统计", self.heading_style))

        total = result.get('total_components', 0)
        passed = result.get('passed_components', 0)
        failed = result.get('failed_components', 0)
        warning = total - passed - failed

        stats_data = [
            [to_para('统计项', self.cell_style), to_para('数量', self.cell_style), to_para('占比', self.cell_style)],
            [to_para('总部件数', self.cell_style), to_para(str(total), self.cell_style), to_para('100%', self.cell_style)],
            [to_para('通过', self.cell_style), to_para(str(passed), self.cell_style), to_para(f'{passed/total*100:.1f}%' if total > 0 else '0%', self.cell_style)],
            [to_para('失败', self.cell_style), to_para(str(failed), self.cell_style), to_para(f'{failed/total*100:.1f}%' if total > 0 else '0%', self.cell_style)],
            [to_para('警告', self.cell_style), to_para(str(warning), self.cell_style), to_para(f'{warning/total*100:.1f}%' if total > 0 else '0%', self.cell_style)],
        ]

        stats_table = Table(stats_data, colWidths=[6*cm, 4*cm, 6*cm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1890ff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f6ffed')),
            ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#f6ffed')),
            ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#fff1f0')),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#fffbe6')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(stats_table)
        elements.append(Spacer(1, 20))

        return elements

    def _create_home_third_comparison(self, result: Dict[str, Any]) -> List:
        """创建首页与第三页比对"""
        elements = []

        comparisons = result.get('home_third_comparison', [])
        if not comparisons:
            return elements

        elements.append(Paragraph("二、首页与第三页字段比对", self.heading_style))

        data = [[to_para('字段名', self.cell_style), to_para('首页值', self.cell_style), to_para('第三页值', self.cell_style), to_para('状态', self.cell_style)]]

        for comp in comparisons:
            field_name = comp.get('field_name', '')
            table_value = comp.get('table_value', '') or '/'
            ocr_value = comp.get('ocr_value', '') or '/'
            is_match = comp.get('is_match', False)

            status = '✓ 一致' if is_match else '✗ 不一致'
            data.append([
                to_para(field_name, self.cell_style),
                to_para(table_value, self.cell_style),
                to_para(ocr_value, self.cell_style),
                to_para(status, self.cell_style)
            ])

        table = Table(data, colWidths=[4*cm, 5*cm, 5*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1890ff')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 20))

        return elements

    def _create_component_details(self, result: Dict[str, Any]) -> List:
        """创建部件核对详情"""
        elements = []

        components = result.get('component_checks', [])
        if not components:
            return elements

        elements.append(Paragraph("三、部件核对详情", self.heading_style))
        elements.append(Paragraph(f"共核对 {len(components)} 个部件", self.note_style))
        elements.append(Spacer(1, 10))

        for idx, item in enumerate(components, 1):
            component_name = item.get('component_name', '未知部件')
            status = item.get('status', 'unknown')
            has_photo = item.get('has_photo', False)
            has_label = item.get('has_chinese_label', False)
            field_comparisons = item.get('field_comparisons', [])
            issues = item.get('issues', [])

            status_colors = {
                'pass': ('通过', colors.HexColor('#52c41a')),
                'fail': ('失败', colors.HexColor('#ff4d4f')),
                'warning': ('警告', colors.HexColor('#faad14')),
            }
            status_text, status_color = status_colors.get(status, ('未知', colors.grey))

            elements.append(Paragraph(f"{idx}. {component_name} [{status_text}]", self.subheading_style))

            photo_status = '✓ 有' if has_photo else '✗ 无'
            label_status = '✓ 有' if has_label else '✗ 无'

            info_data = [
                [to_para('照片覆盖', self.cell_style), to_para(photo_status, self.cell_style)],
                [to_para('中文标签', self.cell_style), to_para(label_status, self.cell_style)],
            ]

            info_table = Table(info_data, colWidths=[3*cm, 4*cm])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(info_table)
            elements.append(Spacer(1, 5))

            if field_comparisons:
                comp_data = [[to_para('字段名', self.cell_style), to_para('表格值', self.cell_style), to_para('OCR值', self.cell_style), to_para('结果', self.cell_style)]]
                for fc in field_comparisons:
                    field = fc.get('field_name', '')
                    table_val = fc.get('table_value', '') or '/'
                    ocr_val = fc.get('ocr_value', '') or '/'
                    match = '✓' if fc.get('is_match') else '✗'
                    comp_data.append([
                        to_para(field, self.cell_style),
                        to_para(table_val, self.cell_style),
                        to_para(ocr_val, self.cell_style),
                        to_para(match, self.cell_style)
                    ])

                comp_table = Table(comp_data, colWidths=[3*cm, 4*cm, 4*cm, 1.5*cm])
                comp_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))

                elements.append(comp_table)
                elements.append(Spacer(1, 5))

            if issues:
                elements.append(Paragraph("问题:", self.note_style))
                for issue in issues:
                    elements.append(Paragraph(f"  • {issue}", self.note_style))
                elements.append(Spacer(1, 5))

            elements.append(Spacer(1, 10))

        return elements

    def _create_issues_summary(self, result: Dict[str, Any]) -> List:
        """创建问题汇总"""
        elements = []

        errors = result.get('errors', [])
        warnings = result.get('warnings', [])

        if not errors and not warnings:
            return elements

        elements.append(Paragraph("四、问题汇总", self.heading_style))

        if errors:
            elements.append(Paragraph("错误:", self.subheading_style))
            for error in errors:
                msg = error.get('message', '')
                elements.append(Paragraph(f"  ✗ {msg}", self.body_style))
            elements.append(Spacer(1, 10))

        if warnings:
            elements.append(Paragraph("警告:", self.subheading_style))
            for warning in warnings:
                msg = warning.get('message', '')
                elements.append(Paragraph(f"  ⚠ {msg}", self.body_style))
            elements.append(Spacer(1, 10))

        return elements

    def _create_footer(self, result: Dict[str, Any]) -> List:
        """创建页脚信息"""
        elements = []

        elements.append(Spacer(1, 30))
        elements.append(Paragraph("— 报告结束 —", ParagraphStyle(
            'Footer',
            fontName=FONT_NAME,
            alignment=TA_CENTER,
            fontSize=9,
            textColor=colors.grey,
        )))
        elements.append(Paragraph(
            f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            self.note_style
        ))

        return elements

    def _fill_excel_overview(self, ws, result: Dict[str, Any]):
        """填充Excel概览sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ws['A1'] = 'PDF报告核对结果'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        ws['A3'] = '文件名'
        ws['B3'] = result.get('filename', '')
        ws['A4'] = '核对时间'
        ws['B4'] = result.get('check_time', '')
        ws['A5'] = '总部件数'
        ws['B5'] = result.get('total_components', 0)
        ws['A6'] = '通过'
        ws['B6'] = result.get('passed_components', 0)
        ws['A7'] = '失败'
        ws['B7'] = result.get('failed_components', 0)

        for cell in ['A3', 'A4', 'A5', 'A6', 'A7']:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

    def _fill_excel_components(self, ws, result: Dict[str, Any]):
        """填充Excel部件核对sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment

        headers = ['序号', '部件名称', '照片', '标签', '状态', '问题']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1890FF', end_color='1890FF', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')

        components = result.get('component_checks', [])
        for idx, item in enumerate(components, 1):
            row = idx + 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item.get('component_name', ''))
            ws.cell(row=row, column=3, value='有' if item.get('has_photo') else '无')
            ws.cell(row=row, column=4, value='有' if item.get('has_chinese_label') else '无')

            status_map = {'pass': '通过', 'fail': '失败', 'warning': '警告'}
            status = status_map.get(item.get('status'), '未知')
            ws.cell(row=row, column=5, value=status)

            issues = item.get('issues', [])
            ws.cell(row=row, column=6, value='; '.join(issues) if issues else '')

            status_colors = {
                'pass': 'C6EFCE',
                'fail': 'FFC7CE',
                'warning': 'FFEB9C',
            }
            color = status_colors.get(item.get('status'), 'FFFFFF')
            ws.cell(row=row, column=5).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    def _fill_excel_issues(self, ws, result: Dict[str, Any]):
        """填充Excel问题汇总sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment

        headers = ['类型', '消息', '页码', '位置']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='1890FF', end_color='1890FF', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')

        row = 2
        for error in result.get('errors', []):
            ws.cell(row=row, column=1, value='错误')
            ws.cell(row=row, column=1).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            ws.cell(row=row, column=2, value=error.get('message', ''))
            ws.cell(row=row, column=3, value=error.get('page_num', ''))
            ws.cell(row=row, column=4, value=error.get('location', ''))
            row += 1

        for warning in result.get('warnings', []):
            ws.cell(row=row, column=1, value='警告')
            ws.cell(row=row, column=1).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            ws.cell(row=row, column=2, value=warning.get('message', ''))
            ws.cell(row=row, column=3, value=warning.get('page_num', ''))
            ws.cell(row=row, column=4, value=warning.get('location', ''))
            row += 1


# 单例
_export_service = None

def get_export_service() -> ReportExportService:
    """获取导出服务单例"""
    global _export_service
    if _export_service is None:
        _export_service = ReportExportService()
    return _export_service
